import sys
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

'''
Analyze xlsx spreadsheet columns for data type and sizes.

Run with `python analyze-xlsx.py <input-file>`

Output is an XLSX file with worksheets that document:
 - name
 - data_type
 - max_length
 - max_value  (e.g. longest str value)
 - original_name

 for each column in the input spreadsheet.

 This is useful for quickly assessing the data types and sizes in a 
 spreadsheet when considering how to best import or convert it.
'''


class SheetInfo:
    def __init__(self, name):
        self.name = name
        self.headers = list()
        self.columns = list()


class FieldInfo:
    def __init__(self, name, data_type, max_length, max_value, original_name):
        self.name = name
        self.data_type = data_type
        self.max_length = max_length
        self.max_value = max_value
        self.original_name = original_name


def clean_name(name):
    non_alpha = '[^A-z0-9 ]'
    SPACE = ' '
    EMPTY = ''
    result = re.sub(non_alpha, SPACE, name)
    result = result.title()
    result = result.replace(SPACE, EMPTY)

    return result


file_name = sys.argv[1]

base = os.path.splitext(file_name)[0]
output_file = base + "-structure.xlsx"

data = list()

input_wb = load_workbook(file_name)

for sheet_name in input_wb.sheetnames:

    input_sheet = input_wb[sheet_name]
    sheet_info = SheetInfo(sheet_name)
    data.append(sheet_info)

    for row_idx, row in enumerate(input_sheet.rows):

        if len(sheet_info.headers) == 0:

            for hdr_idx in range(0, len(row)):
                cell = row[hdr_idx]
                value = cell.value
                if value is None:
                    value = cell.column_letter
                name = clean_name(value)
                sheet_info.headers.append(name)
                sheet_info.columns.append(FieldInfo(
                    name=name,
                    original_name=value,
                    max_length=0,
                    max_value='',
                    data_type=''
                ))

        else:

            for col_idx in range(0, len(row)):

                cell = row[col_idx]

                data_length = 0
                data_type = ''
                data_value = cell.value
                if isinstance(data_value, str):
                    data_length = len(data_value)
                    data_type = 'str'

                if isinstance(data_value, int):
                    data_length = -1
                    data_type = 'int'

                fld_info = sheet_info.columns[col_idx]
                if fld_info.max_length < data_length:
                    fld_info.max_length = data_length
                    fld_info.max_value = data_value

                if data_type != '':
                    if fld_info.data_type == '':
                        fld_info.data_type = data_type
                    elif fld_info.data_type != data_type:
                        fld_info.data_type = 'mixed'

                col_idx += 1


# output results to xlsx file
output_wb = Workbook()

# remove default worksheet
sheet1 = output_wb.active
output_wb.remove(sheet1)

for sheet_info in data:
    output_ws = output_wb.create_sheet(sheet_info.name)

    # output headers
    output_hdrs = ('name', 'data_type', 'max_length', 'max_value', 'original_name')
    bold_font = Font(bold=True, color='000033')
    for idx in range(0, 5):
        hdr_cell = output_ws.cell(row=1, column=idx+1, value=output_hdrs[idx])
        hdr_cell.font = bold_font

    # output field structure analysis
    output_row = 2
    for field_info in sheet_info.columns:

        output_ws.cell(row=output_row, column=1, value=field_info.name)
        output_ws.cell(row=output_row, column=2, value=field_info.data_type)
        output_ws.cell(row=output_row, column=3, value=field_info.max_length)
        output_ws.cell(row=output_row, column=4, value=field_info.max_value)
        output_ws.cell(row=output_row, column=5,
                       value=field_info.original_name)
        output_row += 1

output_wb.save(filename=output_file)
